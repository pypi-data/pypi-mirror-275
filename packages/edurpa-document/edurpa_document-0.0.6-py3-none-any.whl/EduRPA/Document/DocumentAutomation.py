from vietocr.tool.predictor import Predictor
from vietocr.tool.config import Cfg
from PIL import Image
from robot.api.deco import keyword, not_keyword
import cv2
import json
from openpyxl import Workbook
from datetime import datetime
from .transform import perspective_transform, resize_image
import string
import unicodedata
import editdistance
class DocumentAutomation:
    def __init__(self, lang, performance, tolerate, **kwargs):
        if kwargs["dryrun"]:
            # If running dryrun, ignore initialization
            return
        
        self.tolerate = tolerate
        
        cfg = Cfg.load_config_from_name("vgg_seq2seq")
        cfg["device"] = "cpu"
        vgg_seq2seq = Predictor(cfg)

        cfg = Cfg.load_config_from_name("vgg_transformer")
        cfg["device"] = "cpu"
        vgg_transformer = Predictor(cfg)
        
        self.config = {
            "vi": {
              "fast": {
                "predictor": vgg_seq2seq
              },
              "accurate": {
                "predictor": vgg_transformer
              }
            },
            "en": {
                
            }
        }
        self.predictor = self.config[lang][performance]["predictor"]

    @not_keyword
    def image_preprocessing(self, image, template):
        width = template['size']['width']
        height = template['size']['height']
        if template['isScanned'] == True:
            print("Performing resize", template)
            processed_image = resize_image(image, [width, height])
        else:
            print("Performing perspective transformation", template)
            processed_image = perspective_transform(image, [width, height])
        return processed_image   
    
    @not_keyword
    def extract(self, labeled_images):
        results = {}
        for label, image in labeled_images.items():
            print("Predicting label", label)
            prediction = self.predictor.predict(Image.fromarray(image))
            results[label] = prediction

        return results
    
    @keyword("Extract Data From Document", types={'file_name': 'str', 'template': 'dict'})
    def extract_data_from_document(self, file_name, template):
        if type (template) is str:
            template = json.loads(template)
        
        # read and preprocess document (image type)
        document = cv2.imread(file_name)
        document = self.image_preprocessing(document, template)

        # extract images from document
        extracted_labeled_images = {}
        for key in template['dataTemplate'].keys():
            x1 = int(template['dataTemplate'][key]['left'])
            x2 = int(template['dataTemplate'][key]['right'])
            y1 = int(template['dataTemplate'][key]['top'])
            y2 = int(template['dataTemplate'][key]['bottom'])
            extracted_labeled_images[key] = []
            extracted_labeled_images[key] = document[y1:y2, x1:x2]

        # extract data from images
        return self.extract(extracted_labeled_images)
    
    @not_keyword
    def evaluate_answer(self, answer, correct):
        pd = unicodedata.normalize("NFKD", answer).encode("ASCII", "ignore").decode("ASCII")
        gt = unicodedata.normalize("NFKD", correct).encode("ASCII", "ignore").decode("ASCII")
        
        pd = pd.translate(str.maketrans("", "", string.punctuation))
        gt = gt.translate(str.maketrans("", "", string.punctuation))
        
        pd_cer, gt_cer = list(pd.lower()), list(gt.lower())
        dist = editdistance.eval(pd_cer, gt_cer)
        cer = dist / (max(len(pd_cer), len(gt_cer)))
        
        return cer < self.tolerate
    
    # TODO: update this keyword
    @keyword("Create Grade Report File", types={'correct_answer': dict, 'actual_answers': list[dict], 'file_names': list})
    def create_grade_report_file(self, correct_answer, actual_answers, file_names):
        # Initialize report file
        workbook = Workbook()
        report_file_name = f'report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        summarizeSheet = workbook.create_sheet(title=f"Summary")
        header = ['Student Name'] + [label for label in correct_answer.keys()] + ['Total Score']
        summarizeSheet.append(header)

        for (file_name, actual_answer) in zip(file_names, actual_answers):
            summary = []
            lname = file_name.split('.')
            lname.pop()
            lname = '_'.join(lname)
            summary.append(lname)

            sheet = workbook.create_sheet(title=f"Result_{lname}")

            # Write header
            header = ['Label', 'Actual answer', 'Correct answer', 'Score']
            sheet.append(header)

            # Write data rows
            total_score = 0
            for (label , correct_answer_label, actual_answer_label) in zip(correct_answer.keys(), correct_answer.values(), actual_answer.values()):
                is_correct = self.evaluate_answer(actual_answer_label, correct_answer_label)
                row = [label, actual_answer_label, correct_answer_label, is_correct]
                sheet.append(row)
                
                summary.append(actual_answer_label)
                total_score += is_correct

            sheet.append(['','','Total',total_score])
            summary.append(total_score)

            summarizeSheet.append(summary)

        # Remove the default sheet created by openpyxl
        workbook.remove(workbook['Sheet'])

        # Save the Excel file
        workbook.save(report_file_name)

        return report_file_name
    