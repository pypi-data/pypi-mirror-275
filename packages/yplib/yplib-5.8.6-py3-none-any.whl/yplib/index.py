import csv
import hashlib
import json
import os
import platform
import random
import re
import time
import uuid
from datetime import datetime
from datetime import timedelta

import openpyxl
import xlrd

CONFIG_PATH = '.yp.config'

CONFIG_PATH_BAK = 'yp.config'


# 是否是 windows 系统
def is_win():
    return platform.system().lower() == 'windows'


# 是否是 linux 系统, 不是 windows , 就是 linux 系统
def is_linux():
    return not is_win()


# 获得整数类型的 timestamp
def get_timestamp():
    return int(time.time())


# 记录日志, 如果是对象会转化为 json
def to_log(a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
           a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20='', time_prefix=True):
    l = [a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20]
    d = ' '.join(list(map(lambda x: json.dumps(x) if can_use_json(x) else str(x), l)))
    d = d.strip()
    lo = datetime.today().strftime('%Y-%m-%d %H:%M:%S') + ' ' + d if time_prefix else d
    print(lo)
    return lo


# 记录日志, 如果是对象会转化为 json
def to_print(a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
             a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20=''):
    return to_log(a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20,
                  time_prefix=False)


# 将 log 数据, 写入到文件
def to_print_file(a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
                  a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20=''):
    lo = to_print(a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20)
    to_txt(data_list=[lo], file_name=datetime.today().strftime('%Y-%m-%d'), file_path='print', fixed_name=True,
           suffix='.txt')


# 将 log 数据, 写入到文件, 固定文件名称
def to_print_txt(file_name, a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
                 a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20=''):
    lo = to_print(a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20)
    to_txt(data_list=[lo], file_name=file_name, file_path='print', fixed_name=True, suffix='.txt')


# 将 log 数据, 写入到文件
def to_log_file(a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
                a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20='', time_prefix=False):
    lo = to_log(a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20,
                time_prefix=time_prefix)
    to_txt(data_list=[lo], file_name=datetime.today().strftime('%Y-%m-%d'), file_path='log', fixed_name=True,
           suffix='.log')


# 将 log 数据, 写入到固定文件中
def to_log_txt(file_name, a1='tag', a2='', a3='', a4='', a5='', a6='', a7='', a8='', a9='', a10='', a11='', a12='',
               a13='', a14='', a15='', a16='', a17='', a18='', a19='', a20='', time_prefix=True):
    lo = to_log(a1, a2, a3, a4, a5, a6, a7, a8, a9, a10, a11, a12, a13, a14, a15, a16, a17, a18, a19, a20,
                time_prefix=time_prefix)
    to_txt(data_list=[lo], file_name=file_name, file_path='log', fixed_name=True, suffix='.txt')


# 将下划线命名转成驼峰命名
# 例如 : user_id -> userId
# 例如 : USER_ID -> userId
def to_hump_one(s):
    if s is None or s == '':
        return s
    r = ''.join(list(map(lambda x: x.capitalize(), str(s).lower().split('_'))))
    return r[0].lower() + r[1:]


def to_hump(s):
    if s is None or s == '':
        return s
    if isinstance(s, list) or isinstance(s, tuple) or isinstance(s, set):
        return list(map(lambda x: to_hump_one(x), s))
    return to_hump_one(s)


def to_hump_more(a1=None, a2=None, a3=None, a4=None, a5=None):
    r = list(map(lambda x: to_hump(x), [a1, a2, a3, a4, a5]))
    r = list(filter(lambda x: x is not None, r))
    if not len(r):
        return None
    if len(r) == 1:
        return r[0]
    if len(r) == 2:
        return r[0], r[1]
    if len(r) == 3:
        return r[0], r[1], r[2]
    if len(r) == 4:
        return r[0], r[1], r[2], r[3]
    if len(r) == 5:
        return r[0], r[1], r[2], r[3], r[4]


# 将驼峰命名转成下划线命名
# 例如 : userId -> user_id
def to_underline_one(s):
    if s == '' or s is None:
        return s
    return ''.join(list(map(lambda x: '_' + str(x).lower() if x.isupper() else x, str(s))))


def to_underline(s):
    if s == '' or s is None:
        return s
    if isinstance(s, list) or isinstance(s, tuple) or isinstance(s, set):
        return list(map(lambda x: to_underline_one(x), s))
    return to_underline_one(s)


def to_underline_more(a1=None, a2=None, a3=None, a4=None, a5=None):
    r = list(map(lambda x: to_underline(x), [a1, a2, a3, a4, a5]))
    r = list(filter(lambda x: x is not None, r))
    if not len(r):
        return None
    if len(r) == 1:
        return r[0]
    if len(r) == 2:
        return r[0], r[1]
    if len(r) == 3:
        return r[0], r[1], r[2]
    if len(r) == 4:
        return r[0], r[1], r[2], r[3]
    if len(r) == 5:
        return r[0], r[1], r[2], r[3], r[4]


# 是否能用 json
def can_use_json(data):
    if isinstance(data, dict) or isinstance(data, list) or isinstance(data, tuple) or isinstance(data, set):
        return True
    return False


# 检查文件夹是否存在,不存在,就创建新的
# 支持多级目录 , 例如: C:\Users\yangpu\Desktop\study\a\b\c\d\e\f
def check_file(file_name):
    if file_name is None or file_name == '':
        return
    for sep in ['\\', '/']:
        f_n = file_name.split(sep)
        for i in range(1, len(f_n) + 1):
            # C:\Users\yangpu\Desktop\study\p.t
            p_n = sep.join(f_n[0:i])
            if not os.path.exists(p_n):
                os.mkdir(p_n)


# 获得文件名称
# 按照  name_天小时分钟_秒毫秒随机数 的规则来生成
def get_file_name(file_name, suffix='.txt', is_date=False):
    # %Y-%m-%d %H:%M:%S
    [year, month, day, hour, minute, second, ss] = datetime.today().strftime('%Y_%m_%d_%H_%M_%S_%f').split('_')
    s = year + month + day + '_' + hour + minute if is_date else month + day + '_' + hour + minute
    return str(file_name) + '_' + s + '_' + second + random_str(length=4, end_str=52) + suffix


# 文件是否存在
def file_is_empty(file_name=None):
    return file_name is None or file_name == '' or not os.path.exists(file_name)


# md5 算法
def do_md5(data='do_md5'):
    return hashlib.md5(data.encode(encoding='UTF-8')).hexdigest()


# sha256 算法
def do_sha256(data='do_sha256'):
    h = hashlib.sha256()
    h.update(data.encode('utf-8'))
    return h.hexdigest()


# uuid 类型的随机数, 默认 32 位长度
def random_uuid(length=32):
    r = uuid.uuid4().hex
    while len(r) < length:
        r += uuid.uuid4().hex
    return r[0:length]


# 获得随机数
# length    ：随机数长度
# start_str ：随机数开始的字符的位置,从 1 开始, 包含start_str
# end_str   : 随机数结束的字符的位置, 不包含end_str
# 默认的随机数是 : 数字+字母大小写
def random_str(length=64, start_str=1, end_str=62):
    c_s = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_'
    r = ''
    start_str = max(1, start_str)
    end_str = min(len(c_s), end_str)
    while len(r) < length:
        r += c_s[random.Random().randint(start_str, end_str) - 1]
    return r


# 字母的随机数, 默认小写
def random_letter(length=10, is_upper=False):
    r = random_str(length=length, end_str=26)
    return r.upper() if is_upper else r


# 数字的随机数, 返回 int
# 也可以返回指定范围的随机数据
def random_int(length_or_start=10, end=None):
    if end is None:
        return int(random_int_str(length=length_or_start))
    return random.Random().randint(int(length_or_start), int(end) - 1)


# 数字的随机数, 返回 str
def random_int_str(length=10):
    return random_str(length=length, start_str=53, end_str=62)


# 去掉 str 中的 非数字字符, 然后, 再转化为 int
def to_int(s):
    if isinstance(s, int):
        return s
    if s is None or s == '':
        return 0
    if isinstance(s, float):
        return int(s)
    s = re.sub(r'\D', '', str(s))
    # s = ''.join(filter(lambda ch: ch in '0123456789', str(s)))
    # @see https://www.runoob.com/python3/python3-reg-expressions.html
    return 0 if s == '' else int(s)


# 去掉 str 中的 非数字字符, 然后, 再转化为 float
# precision , 小数部位的长度 , 多余的部分 直接去掉
def to_float(s, precision=None):
    if isinstance(s, float):
        s = str(s)
    if s is None or s == '':
        return 0.0
    # s = ''.join(filter(lambda ch: ch in '0123456789.', str(s)))
    # @see https://www.runoob.com/python3/python3-reg-expressions.html
    s = re.sub('[^0-9.]', '', str(s))
    if s == '':
        return 0.0
    if precision is None:
        return float(s)
    s1 = s.split('.')
    return float(s1[0] + '.' + s1[1][0:len(s1[1]) if len(s1[1]) < int(precision) else int(precision)])


# @see https://www.runoob.com/python3/python3-date-time.html
# 将字符串 s 转化成 datetime
def to_datetime(s=None, r_str=False):
    if s is None or s == '':
        return str(datetime.today()) if r_str else datetime.today()
    s = str(s)
    r = None
    date_time_sdf = '%Y-%m-%d %H:%M:%S'
    m_s = {
        "^\\d{4}$": "%Y",
        "^\\d{4}-\\d{1,2}$": "%Y-%m",
        "^\\d{4}-\\d{1,2}-\\d{1,2}$": "%Y-%m-%d",
        "^\\d{4}-\\d{1,2}-\\d{1,2} {1}\\d{1,2}$": "%Y-%m-%d %H",
        "^\\d{4}-\\d{1,2}-\\d{1,2} {1}\\d{1,2}:\\d{1,2}$": "%Y-%m-%d %H:%M",
        "^\\d{4}-\\d{1,2}-\\d{1,2} {1}\\d{1,2}:\\d{1,2}:\\d{1,2}$": date_time_sdf,
        "^\\d{4}-\\d{1,2}-\\d{1,2} {1}\\d{1,2}:\\d{1,2}:\\d{1,2}.\\d{1,9}$": date_time_sdf,
        "^\\d{4}-\\d{1,2}-\\d{1,2}T{1}\\d{1,2}:\\d{1,2}:\\d{1,2}$": date_time_sdf,
        "^\\d{4}-\\d{1,2}-\\d{1,2}T{1}\\d{1,2}:\\d{1,2}:\\d{1,2}.\\d{1,9}$": date_time_sdf,
    }
    for m in m_s:
        if re.match(m, s):
            st = s.split('.')[0]
            st = st.replace('T', ' ')
            r = datetime.strptime(st, m_s[m])
    if r is None and re.match("^\\d{1,13}$", s):
        s_int = int(s)
        if len(s) > 10:
            s_int = int(s_int / 1000)
        time_arr = time.localtime(s_int)
        time_str = time.strftime(date_time_sdf, time_arr)
        r = datetime.strptime(time_str, date_time_sdf)
    if r is None:
        r = datetime.today()
    return str(r) if r_str else r


# 将字符串 s 转化成 datetime, 然后再次转化成 str
def to_datetime_str(s=None):
    return to_datetime(s, r_str=True)


# 时间加减
def to_datetime_add(s=None, days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks=0):
    return to_datetime(s) + timedelta(days=days, seconds=seconds, microseconds=microseconds,
                                      milliseconds=milliseconds, minutes=minutes, hours=hours,
                                      weeks=weeks)


# 将字符串 s 转化成 date 例如: 2021-02-03
def to_date(s=None):
    return str(to_datetime(s))[0:10]


def get_timestamp():
    return int(to_datetime().timestamp())


# 时间加减
def to_date_add(s=None, days=0, seconds=0, microseconds=0, milliseconds=0, minutes=0, hours=0, weeks=0):
    return str(to_datetime_add(s=s, days=days, seconds=seconds, microseconds=microseconds,
                               milliseconds=milliseconds, minutes=minutes, hours=hours, weeks=weeks))[0:10]


# 将 list 中的数据以 json 或者基本类型的形式写入到文件中
# data_list   : 数组数据, 也可以不是数组
# file_name   : 文件名 , 默认 txt
#               当文件名是 C:\Users\yangpu\Desktop\study\abc\d\e\f\a.sql 这种类型的时候, 可以直接创建文件夹,
#                   会赋值 file_name=a,
#                         file_path=C:\Users\yangpu\Desktop\study\abc\d\e\f,
#                         fixed_name=True,
#                         suffix=.sql
#               当文件名是 abc 的时候, 按照正常值,计算
# file_path   : 文件路径
# fixed_name  : 是否固定文件名
# suffix      : 文件后缀, 默认 .txt
# sep_list    : 当 data_list 是 list(list) 类型的时候 使用 sep_list 作为分割内部的分隔符,
#               默认使用 \t 作为分隔符, 如果为 None , 则按照 json 去处理这个 list
def to_txt(data_list,
           file_name='txt',
           file_path='txt',
           fixed_name=False,
           suffix='.txt',
           sep_list='\t',
           file_name_is_date=False):
    file_name = str(file_name)
    for sep in ['\\', '/']:
        f_n = file_name.split(sep)
        if len(f_n) > 1:
            file_name = f_n[-1]
            file_path = sep.join(f_n[0:-1])
            if '.' in file_name:
                suffix = '.' + file_name.split('.')[-1]
                file_name = file_name[0:file_name.rfind('.')]
                fixed_name = True

    # 检查路径 file_path
    while file_path.endswith('/'):
        file_path = file_path[0:-1]
    check_file(file_path)
    # 生成 file_name
    if fixed_name:
        file_name = file_name + suffix
    else:
        file_name = get_file_name(file_name, suffix, is_date=file_name_is_date)
    # 文件路径
    file_name_path = file_name
    if file_path != '':
        file_name_path = file_path + '/' + file_name
    # 写入文件
    text_file = open(file_name_path, 'a', encoding='utf-8')
    if isinstance(data_list, set):
        data_list = list(data_list)
    if not isinstance(data_list, list):
        text_file.write(to_str(data_list) + '\n')
    else:
        for one in data_list:
            if (isinstance(one, list) or isinstance(one, tuple) or isinstance(one, set)) and sep_list is not None:
                text_file.write(str(sep_list).join(list(map(lambda x: to_str(x), one))) + '\n')
            else:
                text_file.write(to_str(one) + '\n')
    text_file.close()
    return file_name_path


# 将 list 中的数据写入到固定的文件中,自己设置文件后缀
def to_txt_file(data_list, file_name=None):
    file_name = datetime.today().strftime('%Y%m%d_%H%M') if file_name is None else file_name
    return to_txt(data_list=data_list, file_name=file_name, file_path='txt', fixed_name=True)


# 转化成字符串
def to_str(data):
    return json.dumps(data, ensure_ascii=False) if can_use_json(data) else str(data)


# 匹配字符串
# 可以参考 正则表达式的操作, 可以使用 chatgpt 帮忙写出这段代码
# pattern = r"CREATE TABLE (\w+)"
# @see https://www.runoob.com/python3/python3-reg-expressions.html
def match_str(pattern, str_a=''):
    match = re.search(pattern, str_a, re.I)
    if match:
        return match.group(1)
    else:
        return None


# # 示例用法
# # 输出：t_admin
# print(match_str(r'create TABLE (\w+)', 'CREATE TABLE t_admin (id bigint(20) NOT NULL'))


# 根据json的key排序,用于签名
# 按照 key 排序, 按照 key=value 然后再 & 连接, 如果数据中有 list, 使用 , 连接 list 中的数据, 然后拼接成 str 返回
# sep       : 分隔符 , 默认 &
# join      : 连接符 , 默认 =
# join_list : list 数据 连接符 , 默认 ,
def sort_by_json_key(data_obj, sep='&', join='=', join_list=','):
    if isinstance(data_obj, list) or isinstance(data_obj, tuple) or isinstance(data_obj, set):
        return join_list.join(list(map(lambda x: f'{x}', data_obj)))
    if not isinstance(data_obj, dict):
        return str(data_obj)
    data_list = sorted(data_obj.items(), key=lambda x: x[0])
    r_l = []
    for one in data_list:
        value_one = one[1]
        if can_use_json(value_one):
            s = sort_by_json_key(data_obj=value_one, sep=sep, join=join, join_list=join_list)
        else:
            s = str(value_one)
        r_l.append(f'{one[0]}{join}{s}')
    return sep.join(r_l)


# 当读取 txt 之类的文件的时候
# 将 txt 文件读取到 list 中, 每一行自动过滤掉行前行后的特殊字符
# sep             : 是否对每一行进行分割,如果存在这个字段,就分割
# sep_all         : 将文件转化成一个字符串,然后对这个字符串,再次总体分割
# start_index     : 从这个地方开始读取,从1开始标号 , 包含这一行
# start_line      : 从这个地方开始读取,从第一行开始找到这个字符串开始标记 , 包含这一行
# end_index       : 读取到这个地方结束,从1开始标号 , 不包含这一行
# end_line        : 读取到这个地方结束,从第一行开始找到这个字符串开始标记 , 不包含这一行
# count           : 读取指定的行数
################################################
# 当读取 excel 之类的文件的时候
# 将 excel 文件读取到 list 中, 可以指定 sheet, 也可以指定列 column_index(列) ,自动过滤掉每个单元格前后的特殊字符
# sheet           : 从 1 开始编号,
# column_index    : 从 1 开始编号, 指定列
# column_index    : 如果是指定值, 这个时候返回的是一个 list, 没有嵌套 list
# column_index    : 如果是 '1,2,3,4'   [1,2,3,4], 返回的是一个嵌套 list[list]
# column_date     : 指定日期格式的列,规则与 column_index 一样
# column_datetime : 指定日期格式的列,规则与 column_index 一样
# 返回的数据一定是一个 list
def to_list(file_name='a.txt',
            sep=None,
            sep_line=None,
            sep_line_contain=None,
            sep_line_prefix=None,
            sep_line_suffix=None,
            sep_all=None,
            ignore_start_with=None,
            ignore_end_with=None,
            start_index=None,
            start_line=None,
            end_index=None,
            end_line=None,
            count=None,
            sheet_index=1,
            column_index=None,
            column_date=None,
            column_datetime=None):
    if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
        return to_list_from_excel(file_name=file_name,
                                  sheet_index=sheet_index,
                                  column_index=column_index,
                                  column_date=column_date,
                                  column_datetime=column_datetime)
    return to_list_from_txt(file_name=file_name,
                            sep=sep,
                            sep_line=sep_line,
                            sep_line_contain=sep_line_contain,
                            sep_line_prefix=sep_line_prefix,
                            sep_line_suffix=sep_line_suffix,
                            sep_all=sep_all,
                            ignore_start_with=ignore_start_with,
                            ignore_end_with=ignore_end_with,
                            start_index=start_index,
                            start_line=start_line,
                            end_index=end_index,
                            end_line=end_line,
                            count=count)


# 当读取 excel 之类的文件的时候
# 将 excel 文件读取到 list 中, 可以指定 sheet, 也可以指定列 column_index(列) ,自动过滤掉每个单元格前后的特殊字符
# sheet_index     : 从 1 开始编号,
# column_index    : 从 1 开始编号, 指定列, 如果是指定值是一个, 这个时候返回的是一个 list, 没有嵌套 list
#                    如果是 '1,2,3,4'   [1,2,3,4], 返回的是一个嵌套 list[list]
# column_date     : 指定日期格式的列,规则与 column_index 一样
# column_datetime : 指定日期格式的列,规则与 column_index 一样
def to_list_from_excel(file_name='a.xls',
                       sheet_index=1,
                       column_index=None,
                       column_date=None,
                       column_datetime=None):
    if file_is_empty(file_name):
        return []
    data_list = list()
    # excel 表格解析成 list 数据
    list_index = []
    for one_index in [column_index, column_date, column_datetime]:
        list_index_one = None
        if one_index is not None:
            list_index_one = []
            if isinstance(one_index, int):
                list_index_one.append(one_index)
            if isinstance(one_index, str):
                i_list = one_index.split(',')
                for i in i_list:
                    list_index_one.append(int(i))
            if isinstance(one_index, list):
                for i in one_index:
                    list_index_one.append(int(i))
        list_index.append(list_index_one)
    list_all = []
    for one_list in list_index:
        if one_list is not None:
            for o in one_list:
                list_all.append(o)
    if len(list_all) > 0 and list_index[0] is not None:
        list_index[0] = list_all
    # 是否是单 list 类型的数据
    list_only_one = False
    if list_index[0] is not None and len(list_index[0]) == 1:
        list_only_one = True
    # 是 xls 格式
    if file_name.endswith('.xls'):
        book = xlrd.open_workbook(file_name)  # 打开一个excel
        sheet = book.sheet_by_index(sheet_index - 1)  # 根据顺序获取sheet
        for i in range(sheet.nrows):  # 0 1 2 3 4 5
            rows = sheet.row_values(i)
            row_data = []
            for j in range(len(rows)):
                cell_data = str(rows[j]).strip()
                is_date = False
                is_datetime = False
                # 日期格式的列
                if list_index[1] is not None and j + 1 in list_index[1]:
                    cell_data = to_date(xlrd.xldate_as_datetime(to_int(rows[j]), 0))
                    is_date = True
                    row_data.append(cell_data)
                    if list_only_one:
                        row_data = cell_data
                # 日期时间格式的列
                if not is_date and list_index[2] is not None and j + 1 in list_index[2]:
                    cell_data = to_datetime(xlrd.xldate_as_datetime(to_int(rows[j]), 0))
                    is_datetime = True
                    row_data.append(cell_data)
                    if list_only_one:
                        row_data = cell_data
                # 指定需要的列
                if not is_date and not is_datetime:
                    if list_index[0] is None:
                        row_data.append(cell_data)
                    else:
                        # 指定需要的列
                        if j + 1 in list_index[0]:
                            row_data.append(cell_data)
                            if list_only_one:
                                row_data = cell_data
            data_list.append(row_data)
    # 是 xlsx 格式
    if file_name.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename=file_name, read_only=True)
        ws = wb[wb.sheetnames[sheet_index - 1]]
        for rows in ws.rows:
            row_data = []
            for j in range(len(rows)):
                cell_data = str(rows[j].value).strip()
                is_date = False
                is_datetime = False
                # 日期格式的列
                if list_index[1] is not None and j + 1 in list_index[1]:
                    cell_data = to_date(cell_data)
                    is_date = True
                    row_data.append(cell_data)
                    if list_only_one:
                        row_data = cell_data
                # 日期时间格式的列
                if not is_date and list_index[2] is not None and j + 1 in list_index[2]:
                    cell_data = to_datetime(cell_data)
                    is_datetime = True
                    row_data.append(cell_data)
                    if list_only_one:
                        row_data = cell_data
                # 指定需要的列
                if not is_date and not is_datetime:
                    if list_index[0] is None:
                        row_data.append(cell_data)
                    else:
                        # 指定需要的列
                        if j + 1 in list_index[0]:
                            row_data.append(cell_data)
                            if list_only_one:
                                row_data = cell_data
            data_list.append(row_data)
    return data_list


# 将一个文件中以空行作为分隔符,
# 组成一个 list(list) 数据
# 多行空行,自动合并到一行空行
def to_list_from_txt_with_blank_line(file_name='a.txt'):
    return to_list_from_txt(file_name, sep_line='')


# 将 list 切分成 list(list)
# 组成一个 list(list) 数据
# 多行空行,自动合并到一行空行
def to_list_list(data_list=[], count=10):
    r_list = []
    o_list = []
    c = 0
    for i in range(len(data_list)):
        o_list.append(data_list[i])
        c += 1
        if c == count:
            r_list.append(o_list)
            o_list = []
            c = 0
    if len(o_list):
        r_list.append(o_list)
    return r_list


# 将一个文件中的数据按照行来区分,
# 会自动过滤掉空格行,
# 组成一个 list(json) 数据
def to_list_json_from_txt(file_name='a.txt',
                          start_index=None,
                          start_line=None,
                          end_index=None,
                          end_line=None,
                          count=None):
    return to_list_from_txt(file_name,
                            start_index=start_index,
                            start_line=start_line,
                            end_index=end_index,
                            end_line=end_line,
                            count=count,
                            line_json=True)


# 将 txt 文件转化成 list 的方法
# 当读取 txt 之类的文件的时候
# 将 txt 文件读取到 list 中, 每一行自动过滤掉行前行后的特殊字符
# sep               : 对每一行进行分割,将 list(str) 转化为 list(list(str)), 或者将 list(list(str)) 转化为 list(list(list(str)))
# sep_line          : 这一行是一个分隔符, 分隔符与这行一样, 将 list(str) 转化为 list(list(str))
# sep_line_contain  : 这一行是一个分隔符,包含这个行分隔符的做分割, 将 list(str) 转化为 list(list(str))
# sep_line_prefix   : 这一行是一个分隔符,以这个分隔符作为前缀的, 将 list(str) 转化为 list(list(str))
# sep_line_suffix   : 这一行是一个分隔符,以这个分隔符作为后缀的, 将 list(str) 转化为 list(list(str))
# sep_is_front      : 这一行，分割行，是包含到前面，还是包含到
# sep_all           : 将文件转化成一个字符串,然后对这个字符串,再次总体分割 将 list(str) 转化为 str , 然后再次转化成 list(str)
# ignore_start_with : 忽略以这个为开头的行
# ignore_end_with   : 忽略以这个为结尾的行
# line_join         : 将 list(list(str)) 转化成 list(str) 类型的数据
# line_json         : 将 list(str) 转化成 list(json) 类型的数据, 会自动过滤掉空格行
# start_index       : 从这个地方开始读取,从1开始标号 , 包含这一行
# start_line        : 从这个地方开始读取,从第一行开始找到这个字符串开始标记 , 包含这一行
# end_index         : 读取到这个地方结束,从1开始标号 , 不包含这一行
# end_line          : 读取到这个地方结束,从第一行开始找到这个字符串开始标记 , 不包含这一行
# count             : 读取指定的行数
def to_list_from_txt(file_name='a.txt',
                     sep=None,
                     sep_line=None,
                     sep_line_contain=None,
                     sep_line_prefix=None,
                     sep_line_suffix=None,
                     sep_is_front=True,
                     sep_all=None,
                     ignore_start_with=None,
                     ignore_end_with=None,
                     line_join=None,
                     line_json=None,
                     start_index=None,
                     start_line=None,
                     end_index=None,
                     end_line=None,
                     count=None):
    if file_is_empty(file_name=file_name):
        return []
    data_list = []
    # 普通文件的解析
    d_list = open(file_name, 'r', encoding='utf-8').readlines()
    # 数量
    c = 0
    start_flag = None
    end_flag = None
    if start_line is not None:
        start_flag = False
    if end_line is not None:
        end_flag = False
    for i in range(len(d_list)):
        line = d_list[i].strip()
        # 判断开始位置
        if start_index is not None and i + 1 < to_int(start_index):
            continue
        # 判断结束位置
        if end_index is not None and i + 1 >= to_int(end_index):
            continue
        # 判断数量
        if count is not None and c >= to_int(count):
            continue
        # 开始标记位
        if start_flag is not None and not start_flag and line.find(start_line) > -1:
            start_flag = True
        # 开始标记位
        if end_flag is not None and not end_flag and line.find(end_line) > -1:
            end_flag = True
        if start_flag is not None and not start_flag:
            # 有开始标记位参数,并且,还没有走到开始标记位
            continue
        elif end_flag is not None and end_flag:
            # 有结束标记位参数,并且,已经走到了结束标记位
            continue
        c += 1
        can_add = True
        if ignore_start_with is not None:
            if isinstance(ignore_start_with, list) or isinstance(ignore_start_with, set):
                for ss in ignore_start_with:
                    if line.startswith(str(ss)):
                        can_add = False
            elif isinstance(ignore_start_with, str):
                if line.startswith(str(ignore_start_with)):
                    can_add = False
        if ignore_end_with is not None:
            if isinstance(ignore_end_with, list) or isinstance(ignore_end_with, set):
                for ss in ignore_end_with:
                    if line.endswith(str(ss)):
                        can_add = False
            elif isinstance(ignore_end_with, str):
                if line.endswith(str(ignore_end_with)):
                    can_add = False
        if can_add:
            data_list.append(line)
    if sep_all is not None:
        # 全部划分, 重新分割成 list(str)
        data_list = ''.join(data_list).split(str(sep_all))
    # 有行分隔符, 将会把 list(str) 转化成 list(list)
    if len(list(filter(lambda x: x is not None, [sep_line, sep_line_prefix, sep_line_contain, sep_line_suffix]))):
        # 当是这种情况的时候,返回的数据结果
        r_list = []
        # 数据中的一行 list 数据
        one_list = []
        for d_o in data_list:
            # 过滤掉空行,无效行
            if len(d_o.strip()) and sep_is_front:
                one_list.append(d_o)
            # 这一行, 等于 sep_line
            if ((sep_line is not None and d_o == sep_line) or
                    # 这一行, 包含 sep_line_contain
                    (sep_line_contain is not None and d_o.find(sep_line_contain) != -1) or
                    # 这一行, 是否是以 sep_line_prefix 开头
                    (sep_line_prefix is not None and d_o.startswith(sep_line_prefix)) or
                    # 这一行, 是否是以 sep_line_suffix 结尾
                    (sep_line_suffix is not None and d_o.endswith(sep_line_suffix))):
                if len(one_list):
                    r_list.append(one_list)
                    one_list = []
            if len(d_o.strip()) and not sep_is_front:
                one_list.append(d_o)
        # 最后的一条数据,兼容一下
        if len(one_list):
            r_list.append(one_list)
        data_list = r_list
    # 对这个 list 进行行内再次分割
    if sep is not None:
        r_list = []
        for line in data_list:
            # list(str) 情况
            if isinstance(line, str):
                r_list.append(line.split(str(sep)))
            # list(list) 情况
            elif isinstance(line, list):
                a_list = []
                for o_line in line:
                    a_list.append(o_line.split(str(sep)))
                r_list.append(a_list)
        data_list = r_list
    # data_list 中的每一个元素都转化成 str
    if line_join is not None:
        data_list = list(map(lambda x: str(line_join).join(x), data_list))
    # data_list 中的每一个元素都转化成 先转化成str, 然后再转化成json
    if line_json is not None and line_json:
        data_list = list(map(lambda x:
                             json.loads(str('' if line_join is None else line_join).join(x)),
                             list(filter(lambda x: x is not None and len(str(x)), data_list))
                             )
                         )
    return data_list


# 读取文件中的数据,返回一个 str
def to_str_from_file(file_name='a.txt',
                     str_join=' ',
                     ignore_start_with=None,
                     ignore_end_with=None,
                     start_index=None,
                     start_line=None,
                     end_index=None,
                     end_line=None,
                     count=None):
    return to_data_from_file(file_name=file_name,
                             ignore_start_with=ignore_start_with,
                             ignore_end_with=ignore_end_with,
                             str_join=str_join,
                             start_index=start_index,
                             start_line=start_line,
                             end_index=end_index,
                             end_line=end_line,
                             count=count,
                             r_str=True)


# 读取文件中的数据,返回一个 json
def to_json_from_file(file_name='a.txt',
                      start_index=None,
                      start_line=None,
                      end_index=None,
                      end_line=None,
                      count=None):
    return to_data_from_file(file_name=file_name,
                             start_index=start_index,
                             start_line=start_line,
                             end_index=end_index,
                             end_line=end_line,
                             count=count,
                             r_json=True)


# 在 to_list 方法上再嵌套一层,
# r_str    : 返回的数据是否是一个 字符串, ''.join(list)
# str_join : 返回的数据是否是一个 字符串, str_join.join(list), 用 str_join 做连接
# r_json   : 返回的数据是否是一个 json 类型的数据
def to_data_from_file(file_name='a.txt',
                      sep=None,
                      sep_line=None,
                      sep_all=None,
                      ignore_start_with=None,
                      ignore_end_with=None,
                      start_index=None,
                      start_line=None,
                      end_index=None,
                      end_line=None,
                      count=None,
                      sheet_index=1,
                      column_index=None,
                      column_date=None,
                      column_datetime=None,
                      r_json=False,
                      str_join='',
                      r_str=False):
    d = to_list(file_name=file_name,
                sep=sep,
                sep_line=sep_line,
                sep_all=sep_all,
                ignore_start_with=ignore_start_with,
                ignore_end_with=ignore_end_with,
                start_index=start_index,
                start_line=start_line,
                end_index=end_index,
                end_line=end_line,
                count=count,
                sheet_index=sheet_index,
                column_index=column_index,
                column_date=column_date,
                column_datetime=column_datetime)
    return str_join.join(d) if r_str else json.loads(str_join.join(d)) if r_json else d


# 将文件导出成excel格式的
def to_excel(data_list, file_name=None, file_path='excel'):
    if file_name is None:
        file_name = 'excel'
    file_name = str(file_name)
    while file_path.endswith('/'):
        file_path = file_path[0:-1]
    check_file(file_path)
    # 实例化对象excel对象
    excel_obj = openpyxl.Workbook()
    # excel 内第一个sheet工作表
    excel_obj_sheet = excel_obj[excel_obj.sheetnames[0]]
    # 给单元格赋值
    for one_data in data_list:
        s_list = []
        if isinstance(one_data, list) or isinstance(one_data, set):
            for one in one_data:
                if isinstance(one, dict) or isinstance(one, list):
                    s = json.dumps(one)
                else:
                    s = str(one)
                s_list.append(s)
            excel_obj_sheet.append(s_list)
        else:
            if can_use_json(one_data):
                s = json.dumps(one_data)
            else:
                s = str(one_data)
            excel_obj_sheet.append([s])

    # 文件保存
    excel_obj.save(file_path + '/' + get_file_name(file_name, '.xlsx', True))


# 将文件导出成csv格式的
# data_list 格式
# data_list = [['Name', 'Age', 'Gender'],
#              ['Alice', 25, 'Female'],
#              ['Bob', 30, 'Male'],
#              ['Charlie', 35, 'Male']]
# data_list = [{
#       "a": 1,
#       "b": 2,
#   },{
#       "a": 1,
#       "b": 2,
# }]
# file_name = 'data'
def to_csv(data_list, file_name=None, file_path='csv'):
    if file_name is None:
        file_name = 'csv'
    file_name = get_file_name(file_name, '.csv', True)
    while file_path.endswith('/'):
        file_path = file_path[0:-1]
    check_file(file_path)
    d_list = []
    if len(data_list) and (isinstance(data_list[0], dict) or isinstance(data_list[0], tuple)):
        title_list = []
        for key in data_list[0]:
            title_list.append(key)
        d_list.append(title_list)
        for one_data in data_list:
            one_list = []
            for k in title_list:
                one_list.append(one_data[k])
            d_list.append(one_list)
    else:
        d_list = data_list
    with open(file_path + '/' + file_name, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(d_list)


# 将数据写入到 config 中
def set_config_data(file_name='config', data=None):
    if data is None:
        data = {}
    set_data_in_user_home(file_name, data)


# 从 config 中获得 配置数据
def get_config_data(file_name='config'):
    # print('get_config_data', file_name)
    config_data = get_data_from_user_home(file_name)
    # print('get_data_from_user_home', config_data)
    if not config_data:
        config_data = get_data_from_path(file_name)
    # print('get_data_from_path', config_data)
    return config_data


# 在当前用户的主目录中, 获得指定文件的数据
def get_data_from_user_home(file_name='config'):
    return get_data_from_path(file_name, os.path.expanduser("~"))


# 将 data 数据,在当前用户的主目录中, 获得指定文件的数据
def set_data_in_user_home(file_name='config', data=None):
    if data is None:
        data = {}
    set_data_in_path(file_name, data, os.path.expanduser("~"))


# 在当前的目录中, 获得指定文件的数据
def get_data_from_path(file_name='config', file_path=None):
    data = get_data_from_path_detail(file_name, file_path, CONFIG_PATH)
    return data if data else get_data_from_path_detail(file_name, file_path, CONFIG_PATH_BAK)


def get_data_from_path_detail(file_name='config', file_path=None, path_name=CONFIG_PATH):
    config_path = file_path + '/' + path_name if file_path else path_name
    # print('config_path_1', config_path)
    if not os.path.exists(config_path):
        # print('config_path_2', config_path)
        return {}
    file_path = config_path + '/' + file_name + '.json'
    # print('config_path_3', file_path)
    if not os.path.exists(file_path):
        return {}
    # print('to_json_from_file', file_path)
    return to_json_from_file(file_path)


# 在当前的目录中, 设置数据到指定路径下
def set_data_in_path(file_name='config', data=None, file_path=''):
    if data is None:
        data = {}
    config_path = file_path + '/' + CONFIG_PATH
    if not os.path.exists(config_path):
        os.mkdir(config_path)
    file_path = config_path + '/' + file_name + '.json'
    text_file = open(file_path, 'w', encoding='utf-8')
    text_file.write(to_str(data))
    text_file.close()
