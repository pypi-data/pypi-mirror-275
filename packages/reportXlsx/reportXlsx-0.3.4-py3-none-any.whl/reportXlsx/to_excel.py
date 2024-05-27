# -*- coding:utf-8 -*- #

import pandas as pd
import time
import ast
from datetime import datetime,timedelta,date
from openpyxl import Workbook as workbook,load_workbook
import xlwings
from copy import copy
import os
from pathlib import Path
from redmail import EmailSender
import configparser
import pkg_resources
from pythonnet import set_runtime
import platform
import inspect

os_type = platform.system()
if os_type == "Windows":
    DLL_FILES = [pkg_resources.resource_filename('reportXlsx', f'resources/windows/{dll_name}') for dll_name in [
        'Spire.XLS.dll','SkiaSharp.dll'
    ]]
    Dll_CLASS = [pkg_resources.resource_filename('reportXlsx',f'resources/windows/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]
else:
    os.environ['PYTHONNET_PYDLL'] = '/usr/share/dotnet/shared/Microsoft.NETCore.App/6.0.30/System.Private.CoreLib.dll'
    os.environ['PYTHONNET_RUNTIME'] = 'coreclr'
    DLL_FILES = [pkg_resources.resource_filename('reportXlsx', f'resources/linux/{dll_name}') for dll_name in [
        'Spire.XLS.dll','SkiaSharp.dll'
    ]]
    Dll_CLASS = [pkg_resources.resource_filename('reportXlsx',f'resources/linux/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]

import clr
clr.AddReference(DLL_FILES[0])
clr.AddReference(DLL_FILES[1])
from openpyxl.styles import numbers
from System.IO import *
from SkiaSharp import *
from Spire.Xls import *
from Spire.Xls.Core.Spreadsheet import HTMLOptions




class EmptyParameterError(Exception):
    """自定义异常类，用于处理参数为空的错误"""
    def __init__(self, parameter_name):
        self.parameter_name = parameter_name
        self.message = f"参数 '{parameter_name}' 不能为空"
        super().__init__(self.message)

class ParameterError(Exception):
    """自定义异常类，用于处理参数类型错误的情况"""
    def __init__(self, message, parameter_name):
        self.parameter_name = parameter_name
        self.message = message
        super().__init__(self.message)

def validate_parameters(params_to_check=None):
    if params_to_check is None:
        params_to_check = []

    def decorator(func):
        def wrapper(*args, **kwargs):
            signature = inspect.signature(func)
            bound_arguments = signature.bind(*args, **kwargs)
            bound_arguments.apply_defaults()
            
            for name, value in bound_arguments.arguments.items():
                if name == 'self':
                    continue
                
                # 只检查指定的参数是否为空
                if name in params_to_check:
                    if value is None or (isinstance(value, str) and not value.strip()):
                        raise EmptyParameterError(name)
                    
                # 参数类型检查
                param = signature.parameters[name]
                if param.annotation is not param.empty and not isinstance(value, param.annotation):
                    raise ParameterError(f"Parameter '{name}' is of type {type(value).__name__}, expected {param.annotation.__name__}", name)
            
            return func(*args, **kwargs)
        return wrapper
    return decorator





class sqlResult_to_Excel:

    def __init__(self,figure_chart=None,rank_chart=None,
                 left_on=1,
                 right_on=1,
                 sorted_col=6,
                 send_date= datetime.now()-timedelta(days=1),
                 pic_dpi = 180,
                 font_path_for_linux=['/usr/share/fonts/chinese'],
                 *args
                 ):
        self.figure_chart = figure_chart #通报前半部分
        self.rank_chart = rank_chart #通报的后半部分
        self.left_on = left_on #左表join的行
        self.right_on = right_on #右表关联的行
        self.sorted_col = sorted_col #排序列
        self.send_date = send_date
        self.pic_dpi = pic_dpi
        self.font_path = font_path_for_linux
    
    def _copy_sheet(self,source_ws, target_ws):
    # 复制数据和样式
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        # 处理合并单元格
        for merge_cell in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merge_cell))

    @validate_parameters(params_to_check=['path'])
    def _solve_cannot_auto_calculate_func(self,path:str)->None:
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open(path)
        excel_book.save()
        excel_book.close()
        excel_app.quit()
    
    @validate_parameters 
    def deal_with_sql_result(self,need_delete_col_list:list)->pd.DataFrame:
        if len(need_delete_col_list) == 0:
            raise EmptyParameterError(need_delete_col_list)
        figure_chart = pd.DataFrame(self.figure_chart)
        rank_chart = pd.DataFrame(self.rank_chart)
        merge_chart = pd.merge(figure_chart,rank_chart,left_on=self.left_on,right_on=self.right_on,how='left').reindex()
        #针对merge_chart排序
        part1 = merge_chart.iloc[:-1].sort_values(by= self.sorted_col,ascending= False)
        last_row = merge_chart.iloc[-1:]
        df_result = pd.concat([part1,last_row],ignore_index=True)
        df_result.columns = list(range(len(df_result.columns)))
        #替换文字
        df_result[1]=df_result[1].apply(self.replace_value)
        df_result.drop(need_delete_col_list,axis= 1 ,inplace = True)
        df_result = df_result.reset_index(drop = True)
        return df_result

    def to_excel(self,
                df_result,
                 excl_start_row = 4,
                 excl_start_col =1,
                 excl_file_path='',
                 **pos):
        wb = load_workbook(excl_file_path)
        sheet_name = self.sheet_name
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title= self.sheet_name)

        start_row = excl_start_row
        start_col = excl_start_col

        for r_idx,row in df_result.iterrows():
            for c_idx,value in enumerate(row):
                ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value)
        wb.save(excl_file_path)

    @validate_parameters(params_to_check=['codilist','metalist','path','sheet_name'])
    def update_by_template(self,codilist:list,metalist:list,path='',dis ='',sheet_name = '')->None:
        wb = load_workbook(path)
        sh_name = sheet_name
        if sh_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title= sheet_name)
        for item1,item2 in zip(codilist,metalist):
            start_row = item1[0]
            start_col = item1[1]
            for r_idx,row in enumerate(item2):
                #print(row),print(type(row))
                for c_idx,value in enumerate(ast.literal_eval(row)):
                    #print(value)
                    ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value)
                    #print(f'row:{start_row+r_idx}',f'column:{start_col+c_idx}')
                    #print(ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value).value)
        if dis:
            wb.save(dis)
        else:
            wb.save(path)
    
    @validate_parameters(params_to_check=['codilist','metalist','path','target_path','sheet_name'])
    def ajust_and_copy_excel(self,source_codilist:list,target_codilist:list,path='',dis = '',sheet_name = ''):
        self._solve_cannot_auto_calculate_func(path)
        wb = load_workbook(path,data_only=True)
        sheet = wb[sheet_name]
        wb.save('test.xlsx')
        #针对所有坐标开展排序工作
        for s_codi,t_codi in zip(source_codilist,target_codilist):
            source_range = sheet[s_codi]
            target = sheet[t_codi]
            start_row = target.row
            start_col = target.column
            for source_row_index,source_row in enumerate(source_range):
                for source_col_index,source_cell in enumerate(source_row):
                    target_row = start_row + source_row_index
                    target_col = start_col + source_col_index
                    #print(target_row,target_col)
                    target_cell = sheet.cell(row=target_row,column=target_col)
                    target_cell.value = source_cell.value
        if dis:
            wb.save(dis)         
        else:
            wb.save(path)

    @validate_parameters(params_to_check=['sort_range_list','sort_col_index_list','path','sheet_name'])
    def sort_excel(self,sort_range_list:list,sort_col_index_list:list,path='',dis = '',if_reverse = True,sheet_name = ''):
        wb = load_workbook(path,data_only=True)
        sheet = wb[sheet_name]
        true_sort_index_list = [x-1 for x in sort_col_index_list]
        for s_range ,sort_index in zip(sort_range_list,true_sort_index_list):
            sort_range = sheet[s_range]
            data = list(sort_range)
            data_all=[]
            for index,row in enumerate(data):
                data_row =[]
                for item in enumerate(row):
                    data_row.append(item[1].value)
                data_all.append(data_row)
            data_all.sort(key = lambda row:row[sort_index],reverse=if_reverse)
        
            start_row = sort_range[0][0].row  # 范围的起始行号
            start_col = sort_range[0][0].column  # 范围的起始列号

            for row_index, row_data in enumerate(data_all):
                for col_index,value in enumerate(row_data):
                    # 计算目标单元格的行和列位置
                    target_row = start_row + row_index
                    target_col = start_col + col_index
                    # 将排序后的值写入目标单元格
                    sheet.cell(row=target_row, column=target_col).value = value
        if dis:
            wb.save(dis)
        else:
            wb.save(path)

    @validate_parameters(params_to_check=['generate_range','path','sheet_name'])
    def excel_to_html(self,generate_range:tuple,path = '',sheet_name = '')->str:
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        # 选 
        cell_range = old_sheet.Range[generate_range] 
        sheet = wb.Worksheets[0]
        cell_range.Copy(sheet.Range[generate_range])
        
        stream = MemoryStream()
        options = HTMLOptions()
        options.SavedAsFragment = True
        options.isExportStyle = True
        sheet.SaveToHtml(stream,options)
        
        stream.Position = 0;
        reader = StreamReader(stream);
        htmlString = reader.ReadToEnd();
        wb.Dispose()
        return htmlString



    @validate_parameters(params_to_check=['starting_col_index','need_delete_cols','path','sheet_name'])
    def del_col_excl(self,starting_col_index:int,need_delete_cols:int,path='',dis= '',sheet_name ='')->None:
        wb =load_workbook(path)
        sh_name = sheet_name
        if sh_name in wb.sheetnames:
            ws = wb[sh_name]
        starting_cols_index = starting_col_index
        num_of_cols_to_delete = need_delete_cols
        ws.delete_cols(starting_cols_index,num_of_cols_to_delete)
        if dis:
            wb.save(dis)
        else:
            wb.save(path)

    @validate_parameters(params_to_check=['results_df_list','name_lists','dis'])
    def add_detail(self,results_df_list:list,name_lists:list,path='',dis = ''):
        if not os.path.exists(path) or not path:
            wb = workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
        else:
            wb = load_workbook(path)
        for item,name in zip(results_df_list,name_lists):
            new_sheet = wb.create_sheet(title=name)
            for row in item:
                str_row = tuple(str(item) for item in row)
                new_sheet.append(str_row)
        if not dis:
            wb.save(path)
        else:
            wb.save(dis)

    @validate_parameters(params_to_check=['sheet_need_change_list','need_set_col_tuple','path'])
    #设置查看参数
    def set_up_style_of_the_workbook(self,sheet_need_change_list:list,need_set_col_tuple:list,path='',dis = ''):
        wb = load_workbook(path)
        for sheet,need_row in zip(sheet_need_change_list,need_set_col_tuple):
            ws = wb[sheet]  
            max_row = ws.max_row
            max_col = ws.max_column

            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    text_value = cell.value
                    try:
                        numeric_value = float(text_value)
                        cell.value = numeric_value
                    except (ValueError, TypeError):
                        pass
            print(need_row)
            for column in need_row:
                print(need_row)
                for row in range(1, max_row + 1):
                    cell = ws.cell(row=row, column=column)
                    cell.number_format = numbers.FORMAT_PERCENTAGE
        if dis:
            wb.save(dis)
        else:
            wb.save(path)

    @validate_parameters(params_to_check=['target_excel','target_sheet'])
    def split_the_excel(self,target_excel:str,column_index = 1,target_sheet =''):
        wb = load_workbook(target_excel)
        sheet = wb[target_sheet]
        ws = wb.active
        # 创建一个字典，用于存储按照字段值拆分后的数据
        split_data = {}
        #遍历每一行，根据指定字段的值进行拆分
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[column_index - 1]  # 由于列索引从1开始，所以需要减1
            if key not in split_data:
                split_data[key] = []
            split_data[key].append(row)

        # 根据拆分后的数据创建新的Excel文件
        for key, data in split_data.items():
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.append(ws[1])  # 复制原始Excel的标题行
            for row in data:
                ws_new.append(row)
            wb_new.save(f"{key}.xlsx")
    
    @validate_parameters(params_to_check=['target_wb','need_add_workbook','need_sheet_name'])
    def add_sheet(self,target_wb='',need_add_workbook='',need_sheet_name='',new_sheet_name='new sheet'):
        #添加sheet
        wb = load_workbook(target_wb)
        wb_old = load_workbook(need_add_workbook)
        ws_old = wb_old[need_sheet_name]
        ws_new = wb.create_sheet(new_sheet_name)

        self._copy_sheet(ws_old,ws_new)
        wb.save(target_wb)
    
    @validate_parameters(params_to_check=['target_wb','source_workbook','source_sheet_name'])
    def add_sheet_with_style(self,target_wb='',source_workbook='',source_sheet_name='',new_sheet_name='new sheet'):
        #添加sheet
        wb = Workbook()
        wb.LoadFromFile(target_wb)
        if wb.Worksheets[new_sheet_name]:
            wb.Worksheets.Remove(wb.Worksheets[new_sheet_name])
        source_wb = Workbook()
        source_wb.LoadFromFile(source_workbook)
        old_sheet = source_wb.Worksheets[source_sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)

        wb.Worksheets[source_sheet_name].Name = new_sheet_name
        new_sheet = wb.Worksheets[new_sheet_name]
        wb.ActiveSheetIndex = wb.Worksheets.Count - 1
        wb.SaveToFile(target_wb,FileFormat.Version2013)
    
    @validate_parameters(params_to_check=['to_pic_generate_range','path'])
    def excel_to_pic(self,to_pic_generate_range,path = '',sheet_name = '通报',pic_name = 'results.png'):
        #生成新xlsx
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)
        wb.Worksheets.RemoveAt(0)
        wb.Worksheets.RemoveAt(1)
       
        setting = ConverterSetting()
        setting.XDpi = self.pic_dpi
        setting.YDpi = self.pic_dpi
        wb.ConverterSetting =setting
        sheet = wb.Worksheets[sheet_name]
        
        os_type = platform.system()
        if os_type == "Windows":
            sheet.SaveToImage(pic_name,to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3] )
            wb.Dispose()
        elif os_type == "Linux":
            wb.CustomFontFileDirectory= self.font_path
            image_stream = sheet.ToImage(to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3])
            managedStream = SKManagedStream(image_stream)
            bitmap = SKBitmap.Decode(managedStream)
            image = SKImage.FromBitmap(bitmap)
            data = image.Encode(SKEncodedImageFormat.Png,dpi)
            File.WriteAllBytes(pic_name,data.ToArray())
            wb.Dispose()
        else:
            print("不支持当前系统")

if __name__ =='__main__':
    pass