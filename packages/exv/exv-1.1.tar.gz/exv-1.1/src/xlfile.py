import openpyxl as xlx
import xlrd

class xlFileErrorSheetnameError(Exception):
    pass


class xlFile:
    '''
    Convenience class for supporting .xlsx and .xls files
    without caring about details.
    '''

    def __init__(self, filename):
        self.wb = None

    @classmethod
    def load_excel_file(cls, filename):
        try:                    # Most of the time we will get .xlsx files, so let's try that
            return xlsxFile(filename)
        except xlx.utils.exceptions.InvalidFileException:
            # Try open an .xls file
            return xlsFile(filename) # Need exception handling here too?


class xlsxFile:
    def __init__(self, filename):
        self.wb = xlx.load_workbook(filename, data_only=True)

    def sheet_names(self):
        return self.wb.sheetnames


    def sheet(self, sheetname):
        try:
            return xlsxSheet(self.wb[sheetname])
        except KeyError:
            raise xlFileErrorSheetnameError()
            


class xlsxSheet:
    def __init__(self, sheet):
        self.sh = sheet

    def n_columns(self):
        return self.sh.max_column

    def rows(self):
        return self.sh.values



class xlsFile:
    def __init__(self, filename):
        self.wb = xlrd.open_workbook(filename)

    def sheet_names(self):
        return self.wb.sheet_names()

    def sheet(self, sheetname):
        return xlsSheet(self.wb.sheet_by_name(sheetname))


class xlsSheet:
    def __init__(self, sheet):
        self.sh = sheet

    def n_columns(self):
        return self.sh.ncols

    def rows(self):
        for r in self.sh.get_rows():
            yield map(lambda c: c.value, r)

    
