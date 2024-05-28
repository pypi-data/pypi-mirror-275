import argparse
from tabulate import tabulate, tabulate_formats
from openpyxl import load_workbook
from version import __version__
import string

def setup_argparse():
    ap = argparse.ArgumentParser(prog='exv', description='Command line tool for viewing Excel files.') 
    ap.add_argument('infile', help='An Excel file')
    ap.add_argument('worksheet', nargs='?',
                    help='View the named worksheet. Without a worksheet either the single existing worksheet is viewed or a list of worksheets is displayed.')
    ap.add_argument('-n', '--no-indices', action='store_true',
                    help='Do not show row and column indices.')
    ap.add_argument('-nc', '--no-col-indices', action='store_true',
                    help='Do not show column indices.')
    ap.add_argument('-nr', '--no-row-numbers', action='store_true',
                    help='Do not show row numbers.')
    ap.add_argument('-f', '--format', choices=tabulate_formats,
                    default='plain',
                    help='Output format.')
    ap.add_argument('-v', '--version', action='version', version='%(prog)s '+__version__)    

    return ap


def columns_ident(i):
    '''
    Map integer i to corresponding Excel column name. 
    0 -> A
    1 -> B
    ...
    25 -> Z
    26 -> AA
    27 -> AB
    ...
    '''
    if i > 25:
        prefix = columns_ident(i // 26 - 1)
    else:
        prefix = ''
    return prefix + (string.ascii_uppercase[i % 26])


def view_worksheet(book, sheetname, args):
    sh = book[sheetname]
    tab_rows = []

    for row in sh.values:
        new_row = []
        for cell in row:
            new_row.append(cell)
        tab_rows.append(new_row)

    header = list(map(columns_ident, range(sh.max_column)))
    if args.no_indices:
        print(tabulate(tab_rows, tablefmt=args.format))
    elif args.no_col_indices:
        print(tabulate(tab_rows, showindex='always', tablefmt=args.format))
    elif args.no_row_numbers:
        print(tabulate(tab_rows, headers=header, tablefmt=args.format))
    else:
        print(tabulate(tab_rows, headers=header, showindex='always', tablefmt=args.format))

    

def main():
    ap = setup_argparse()
    args = ap.parse_args()

    wb = load_workbook(args.infile, data_only=True)
    if args.worksheet:
        view_worksheet(wb, args.worksheet, args)
    elif len(wb.sheetnames) > 1:
        print('Available sheets:')
        for sheetname in wb.sheetnames:
            print(sheetname)
    else:
        view_worksheet(wb, wb.sheetnames[0], args)
        

if __name__ == '__main__':
    main()
