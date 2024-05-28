import argparse
from tabulate import tabulate, tabulate_formats
#import xlrd
from openpyxl import load_workbook
from version import __version__

def setup_argparse():
    ap = argparse.ArgumentParser(prog='exv', description='Command line tool for viewing Excel files.') 
    ap.add_argument('infile', help='An Excel file')
    ap.add_argument('worksheet', nargs='?',
                    help='View the named worksheet. Without a worksheet either the single existing worksheet is viewed or a list of worksheets is displayed.')
    ap.add_argument('-nr', '--no-row-numbers', action='store_true')
    ap.add_argument('-f', '--format', choices=tabulate_formats,
                    default='plain',
                    help='Output format.')
    ap.add_argument('-v', '--version', action='version', version='%(prog)s '+__version__)    

    return ap


def view_worksheet(book, sheetname, args):
    sh = book.get_sheet_by_name(sheetname)
    tab_rows = []
    for r, row in enumerate(sh.values):
        if args.no_row_numbers:
            new_row = []
        else:
            new_row = [r]
        for cell in row:
            new_row.append(cell)
        tab_rows.append(new_row)
    print(tabulate(tab_rows, tablefmt=args.format))

def main():
    ap = setup_argparse()
    args = ap.parse_args()

    wb = load_workbook(args.infile, data_only=True)
    if args.worksheet:
        view_worksheet(wb, args.worksheet, args)
    elif len(wb.sheetnames) > 1:
        print('Available sheets:')
        for sheetname in wb.sheetnames:
            print(sheetname)
    else:
        view_worksheet(wb, wb.sheetnames[0], args)
        

if __name__ == '__main__':
    main()
