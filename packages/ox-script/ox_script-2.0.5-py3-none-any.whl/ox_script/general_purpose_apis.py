# MIT License

# Copyright (c) 2023 Postek Electronics Co., Ltd

# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

# For the functions that only have pass in this file, please refer to the ox script development manual for details
# The functions are marked with pass only because they are meanted to be executed on the printer and will not function
# correctly on your device. Please use the runtime environment on Postek printers for most accurate results

# ===========================================================#
# General purpose functions
# These functions can be executed directly on your device
# and is designed to help you with data parsing, connecting to
# databases, and getting data in the format to be printed
# ===========================================================#

from pandas import DataFrame
import openpyxl
import os
import re

# This class defines the basis of a database to be used like this
# data_base = PTKDataBase()
# data_base.read_excel_file("example/path/example_file.xlsx")


class PTKDataBase:
    def __init__(self):
        self.data = {}

    def __getitem__(self, key):
        return self.data[key]

    def __repr__(self) -> str:
        return self.__str__()

    def keys(self):
        return self.data.keys()

    def __str__(self) -> str:
        return self.data.__str__()

    def pprint(self, col_spacing=10) -> str:
        returnStr = "{"
        for key in self.data:
            returnStr += "\n"
            returnStr += f"  Table Name : {key} \n"
            lines = DataFrame(
                self.data[key].data_table).to_string(col_space=col_spacing)
            for line in lines.split("\n"):
                returnStr += f"\n\t{line}"
            returnStr += "\n\n"
            returnStr += "------------------------------------------------------------"
            returnStr += "\n"
        returnStr = returnStr.rsplit(
            "------------------------------------------------------------", 1)[0] + "}"
        print(returnStr)

    def read_excel_file(self, file_path: str):
        ws = openpyxl.load_workbook(file_path).active
        column_names = [cell.value for cell in ws[1]]
        base_path = os.path.splitext(os.path.basename(file_path))[0]
        temp = {}
        for column_index, column_name in enumerate(column_names):
            column, column_name = [], ""
            for cell in ws.iter_rows(
                min_row=1, min_col=column_index + 1, max_col=column_index + 1
            ):
                if cell[0].value == None:
                    continue
                if column_name == "":
                    column_name = cell[0].value
                else:
                    column.append(str(cell[0].value).strip())
            if column_name != "" and len(column) != 0:
                temp[column_name] = column
        self.data[base_path] = DataBaseEntry(temp)

# This class defines the basis for an entry in the data base, not to be used directly
class DataBaseEntry:
    def __init__(self, data_table: map):
        self.data_table = data_table

    def __str__(self) -> str:
        return self.data_table.__str__()

    def __repr__(self) -> str:
        return self.__str__()

    def __getitem__(self, key):
        return self.data_table[key]

    def get_max_row(self):
        print(self.data_table.keys())
        max_row = 0
        for key in self.data_table.keys():
            if len(self.data_table[key]) > max_row:
                max_row = len(self.data_table[key])
        return max_row

    def get_max_col(self):
        return len(self.data_table.keys())

    def get_column_headers(self):
        return self.column_headers


def retrieve_data(item_list: list, index: int, excluding=""):
    """
    retrieve_data is used to retrieve data from a list of items. It can be used to retrieve data from a list
        of items and user can use the excluding parameter to exclude items from the list. the excluding parameter
        should be in the following format "starting_index;ending_index;exclusion_index". 

    Parameters:
        item_list (list): The list of items to be retrieved from   
        index (int): The index of the item to be retrieved
        excluding (str, optional): The string that defines the exclusion of items. Defaults to "".

    Returns:
        str: The data from the item_list at the index specified with the exclusion applied
    """
    if excluding == "":
        return item_list[index]
    else:
        pattern = r'^([\d,-]*;){2}[\d,-]*$'
        if re.match(pattern=pattern, string=excluding):
            return _exclusion(item_list, index, excluding)
        else:
            print(
                "excluding string in wrong format, it should follow starting_index;ending_index;exclusion_index(, and - allowed)")
            return "exclusion string error"

# This is an internal function. Not meant to be used directly


def _exclusion(item_list: list, current_index: int, parameters=";;"):
    parameters = parameters.split(";")
    i = 0
    for items in parameters:
        if items == "":
            if i == 0:
                parameters[0] = "1"
            elif i == 1:
                parameters[1] = str(len(item_list))
            elif i == 2:
                parameters[2] = ""
        i = i + 1
    temp = item_list[:]
    return _get_item_with_exclusion(
        temp, current_index, int(parameters[0]), int(
            parameters[1]), parameters[2]
    )

# This is an internal function. Not meant to be used directly


def _get_item_with_exclusion(
    item_list: list,
    current_index: int,
    starting_index: int,
    ending_index: int,
    exclusion_index: str,
):
    try:
        current_index = current_index + starting_index
        if (current_index > ending_index):
            return "Index Reached Ending Index"
        excluding_rows = []
        if exclusion_index != "":
            if "," in exclusion_index:
                exclusion_index = exclusion_index.split(",")
                for item in exclusion_index:
                    excluding_rows.extend(_get_excluding_row(item))
            else:
                excluding_rows = _get_excluding_row(exclusion_index)
        num_removed = 0
        for item in excluding_rows:
            if (int(item) > ending_index) or int(item) > len(item_list):
                print("exclusion index greater than ending index")
                break
            item_list.remove(item_list[int(item) - num_removed])
            num_removed = num_removed + 1
        return item_list[current_index]
    except IndexError:
        return "Index Error"

# This is an internal function. Not meant to be used directly


def _get_excluding_row(exclusion_index: str):
    if "-" in exclusion_index:
        exclusion_index = exclusion_index.split("-")
        return [
            str(num)
            for num in range(int(exclusion_index[0]), int(exclusion_index[1]) + 1)
        ]
    else:
        return [(exclusion_index)]

# ===========================================================#
# Functions related to changing printjob forms genereated by
# label editing softwares so the label can be easily designed
# in other editing software and Ox Script can be used to update
# the data on the label
# ===========================================================#


basepath = ""


def PTK_GetAllFormVariables(filepath) -> dict:
    """
    PTK_GetAllFormVariables is used to retrive all the variables in the form so they can be
        updated with the PTK_UpdateAllFormVariables function.

    Parameters:
        filepath (string): The path to the form file. The form file can be generated by
            label editing softwares and just place the form file next to the script so
            the relative path can be accessed

    Returns:
        dict: A dictionary of the positions of the variables are in the form so that it can be
            used to update the variables in the form
    """
    variable_locations = {}
    if os.access(filepath, os.F_OK):
        with open(filepath, "r") as fp:
            # read all lines using readline()
            lines = fp.readlines()
            for line in lines:
                # check if string present on a current line
                word = "#OX:"
                if line.find(word) != -1:
                    for item in reversed(line.split(",")):
                        if item.find(word) != -1:
                            key = item
                            item = item.replace(word, "")
                            item = item.replace('"', "")
                            item = item.replace("#\n", "")
                            if item not in variable_locations.keys():
                                variable_locations[item] = [
                                    {
                                        "position": lines.index(line),
                                        "line": line,
                                        "replacement_key": key,
                                    },
                                ]
                            elif item in variable_locations.keys():
                                variable_locations[item].append(
                                    {
                                        "position": lines.index(line),
                                        "line": line,
                                        "replacement_key": key,
                                    }
                                )
        return variable_locations


def PTK_UpdateAllFormVariables(filename, **kwargs) -> str:
    """
    PTK_UpdateAllFormVariables is used to update all the variables in the form

    Parameters:
        filename (string): The name of the form file. The form file can be generaed by
            label editing softwares and just place the form file next to the script so
            the relative path can be accessed
        **kwargs: The key value pair of the variables to be updated

    e.x.:
        cmd = PTK_UpdateAllFormVariables(
                'command1.txt',
                Input1=data[0],
                Input2=data[1],
            )
        PTK_SendCmdToPrinter(cmd)
        where 'command1.txt' is the form file name and 'Input1' and 'Input2' are the variables
            name defined when the form file is generated

    Returns:
        str: The updated form file in string that can be sent to the printer through the function
            PTK_SendCmdToPrinter
    """
    filepath = basepath + filename
    if os.access(filepath, os.F_OK):
        variable_locations = PTK_GetAllFormVariables()(filepath)
        with open(filepath, "r") as text_file:
            lines = [line for line in text_file]
            for key, value in kwargs.items():
                if key in variable_locations.keys():
                    for items in variable_locations[key]:
                        position = items["position"]
                        line = items["line"]
                        line = line.replace(
                            items["replacement_key"], '"' + value + '"')
                        lines[position] = line + "\n"
            return "".join(lines) + "\r\n"
    else:
        print("Form file specified does not exist")
        return ""
